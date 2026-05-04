"""
Retry script: Only re-runs rows that have 'UI Error' status.
Uses longer timeouts and waits to handle slow website responses.
"""
from playwright.sync_api import sync_playwright
import time
import re
import sys
import io
import openpyxl
from openpyxl.cell.cell import MergedCell
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_PATH = Path(__file__).parent / "IT23827080_Assignment 1 - Test cases.xlsx"
URL = "https://www.pixelssuite.com/chat-translator"
WAIT_MS = 9000
RETRIES = 15
RETRY_WAIT_MS = 2000
TIMEOUT_MS = 120000
TYPE_DELAY_MS = 40

INPUT_COL = 3
EXPECTED_COL = 4
ACTUAL_COL = 5
STATUS_COL = 6


def _merged_top_left(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)


def _dismiss_overlays(page):
    for role, name in [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it)$", re.IGNORECASE)),
        ("button", re.compile(r"^(Accept all|Accept All)$", re.IGNORECASE)),
    ]:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(500)
        except Exception:
            pass


def _find_locators(page):
    deadline = time.time() + (TIMEOUT_MS / 1000)
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            inp = page.locator('textarea[placeholder*="English"]').first
            out = page.locator('textarea[placeholder*="Sinhala"]').first
            if inp.count() > 0 and out.count() > 0 and inp.is_visible() and out.is_visible():
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return inp, out, btn
        except Exception:
            pass
        try:
            textareas = [page.locator("textarea").nth(i) for i in range(page.locator("textarea").count())]
            visible = [t for t in textareas if t.is_visible()]
            if len(visible) >= 2:
                btn = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return visible[0], visible[1], btn
        except Exception:
            pass
        page.wait_for_timeout(500)
    raise RuntimeError("Could not find input/output textareas on page.")


def _clear_and_type(page, locator, text):
    for _ in range(3):
        try:
            locator.click(timeout=3000)
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        except Exception:
            pass
        try:
            locator.fill("")
        except Exception:
            pass
    try:
        locator.click(timeout=3000)
        locator.type(text, delay=TYPE_DELAY_MS)
    except Exception:
        locator.fill(text)


def _read_output(locator):
    for method in ["input_value", "inner_text", "text_content"]:
        try:
            v = getattr(locator, method)()
            if v and str(v).strip():
                return str(v).strip()
        except Exception:
            pass
    return ""


def main():
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb.active

    # Find rows with UI Error in STATUS_COL that have input in INPUT_COL
    retry_rows = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=STATUS_COL)
        if cell.value == "UI Error":
            input_cell = _merged_top_left(ws, r, INPUT_COL)
            if input_cell.value and str(input_cell.value).strip():
                retry_rows.append(r)

    if not retry_rows:
        print("No rows with 'UI Error' status found. All tests are already resolved!")
        return

    print(f"Found {len(retry_rows)} rows to retry: {retry_rows}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, slow_mo=50)
        page = browser.new_page()
        page.set_default_timeout(TIMEOUT_MS)

        print(f"Opening {URL}...")
        page.goto(URL, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=TIMEOUT_MS)
        except Exception:
            pass
        page.wait_for_selector("textarea", timeout=TIMEOUT_MS)
        print("Page loaded.")

        input_loc, output_loc, action_loc = _find_locators(page)
        print("Found UI elements.")

        for row in retry_rows:
            input_cell = _merged_top_left(ws, row, INPUT_COL)
            singlish = str(input_cell.value).strip()
            expected_cell = _merged_top_left(ws, row, EXPECTED_COL)
            expected = str(expected_cell.value).strip() if expected_cell.value else ""

            print(f"\n[Row {row}] Testing: {singlish}")
            try:
                _dismiss_overlays(page)
                prev = _read_output(output_loc)

                _clear_and_type(page, input_loc, singlish)

                if action_loc:
                    try:
                        action_loc.click(timeout=5000)
                    except Exception:
                        pass

                page.wait_for_timeout(WAIT_MS)

                actual = ""
                for attempt in range(RETRIES):
                    cur = _read_output(output_loc)
                    if cur and cur != prev:
                        actual = cur
                        break
                    print(f"  Attempt {attempt+1}: waiting...")
                    page.wait_for_timeout(RETRY_WAIT_MS)

                if not actual and prev:
                    # Try clicking Transliterate again
                    print("  Retrying click...")
                    try:
                        action_loc.click(timeout=5000)
                    except Exception:
                        pass
                    page.wait_for_timeout(WAIT_MS)
                    for attempt in range(RETRIES):
                        cur = _read_output(output_loc)
                        if cur and cur != prev:
                            actual = cur
                            break
                        page.wait_for_timeout(RETRY_WAIT_MS)

                if not actual:
                    actual = _read_output(output_loc) or ""

                ws.cell(row=row, column=ACTUAL_COL).value = actual

                if expected:
                    status = "PASS" if actual == expected else "FAIL"
                else:
                    status = "COLLECTED"
                ws.cell(row=row, column=STATUS_COL).value = status
                print(f"  -> Actual: {actual[:60]!r}")
                print(f"  -> Status: {status}")

                wb.save(str(EXCEL_PATH))

            except Exception as e:
                print(f"  ERROR: {e}")
                ws.cell(row=row, column=STATUS_COL).value = "UI Error"
                wb.save(str(EXCEL_PATH))

        browser.close()

    wb.save(str(EXCEL_PATH))
    print("\nDone! Results saved.")

    # Final summary
    wb2 = openpyxl.load_workbook(str(EXCEL_PATH))
    ws2 = wb2.active
    counts = {}
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(row=r, column=INPUT_COL).value:
            s = ws2.cell(row=r, column=STATUS_COL).value
            counts[s] = counts.get(s, 0) + 1
    print("\nFinal Summary:")
    for k, v in sorted(counts.items()):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
